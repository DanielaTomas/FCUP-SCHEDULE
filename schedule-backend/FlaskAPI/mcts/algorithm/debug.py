from algorithm.macros import WAIT_TIME, MAX_RETRIES
from algorithm.simulation_results_writer import directory_exists
import graphviz
import plotly.graph_objs as go
from plotly.subplots import make_subplots
import os
import openpyxl
from openpyxl.styles import Font
import re
import time
import io, pstats

def visualize_tree(root, output_file_name = "mcts_tree"):
    """
    Creates a MCTS tree using Graphviz and saves it as a PDF.

    Parameters:
        root (Node): Root node of the MCTS tree.
        output_file_name (str): Filename for the output PDF.
    """
    print("Processing the tree...")
    dot = graphviz.Digraph(comment = 'MCTS Tree')

    def add_nodes_edges(node):
        if not node: return
        label = f"{node.assignment} {node.best_hard_penalty} {node.best_soft_penalty}" #{node.expansion_limit}"
        dot.node(str(id(node)), label=label, shape="plaintext", width="0.01", height="0.01")
        #label = ""
        #dot.node(str(id(node)), label=label, shape="point", width="0.01", height="0.01")

        for child in node.children:
            if child:
                dot.edge(str(id(node)), str(id(child)), dir="none", style="solid", penwidth="0.5")
                add_nodes_edges(child)

    try:
        add_nodes_edges(root)

        output_dir = "mcts_tree"
        directory_exists(output_dir)
        output_path = os.path.join(output_dir, output_file_name)

        dot.render(output_path, format="pdf", cleanup=True)
        print(f"Tree successfully saved as {output_path}.pdf\n")

    except graphviz.ExecutableNotFound:
        print("Error: Graphviz is not installed or not in the system PATH.")
    except graphviz.CalledProcessError:
        print("Error: Graphviz failed to render the file. The tree may be too large.")
    except MemoryError:
        print("Error: Not enough memory to render the graph.")
    except Exception as e:
        print(f"Unexpected error: {e}")


def plot_progress(metrics, output_file_name = "constraint_progress.html"):
    """
    Plots the progress of hard and soft constraints over iterations.

    Parameters:
        metrics (dict): Dictionary with keys: 'iterations', 'current_hard', 'best_hard', 'current_soft', 'best_soft'.
        output_file_name (str): Filename for the saved HTML plot.
    """
    print("Processing the constraint progress...")
    fig = make_subplots(rows=1, cols=2, 
                        subplot_titles=("Hard Constraint Progress", "Soft Constraint Progress"),
                        shared_xaxes=True)

    fig.add_trace(go.Scatter(x=metrics["iterations"], y=metrics["current_hard"], mode='lines+markers', name='Current Hard', line=dict(color='blue')), row=1, col=1)
    fig.add_trace(go.Scatter(x=metrics["iterations"], y=metrics["best_hard"], mode='lines+markers', name='Best Hard', line=dict(color='green')), row=1, col=1)
    
    fig.add_trace(go.Scatter(x=metrics["iterations"], y=metrics["current_soft"], mode='lines+markers', name='Current Soft', line=dict(color='red')), row=1, col=2)
    fig.add_trace(go.Scatter(x=metrics["iterations"], y=metrics["best_soft"], mode='lines+markers', name='Best Soft', line=dict(color='purple')), row=1, col=2)

    fig.update_layout(title="Constraint Progress (Current vs. Best)", showlegend=True)
    fig.update_xaxes(title_text="Iteration")
    fig.update_yaxes(title_text="Constraint Value")
    
    output_dir = "constraint_progress"
    directory_exists(output_dir)
    output_file_name = os.path.join(output_dir, output_file_name)

    fig.write_html(output_file_name)
    #fig.show()
    print(f"Plot saved successfully to {output_file_name}\n")


def save_results_to_excel(results, file_names, filename="test_results.xlsx"):
    """
    Saves test results to an Excel file with a new sheet for each test run.

    Parameters:
        results (dict): Dictionary mapping filenames to lists of [time, hard, soft] results.
        file_names (list): Ordered list of filenames used for the results.
        filename (str): Path to the Excel file to save results.
    """
    print(f"Saving results to excel...")
    
    for attempt in range(MAX_RETRIES):
        try:
            if os.path.exists(filename):
                wb = openpyxl.load_workbook(filename)
                sheet_name = f"Test {len(wb.sheetnames) + 1}"
                ws = wb.create_sheet(title=sheet_name)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Test 1"
            
            header = ["TEST"] + file_names
            ws.append(header)

            def get_value(file, index):
                return results[file][index] if file in results and len(results[file]) > index else "N/A"
            
            best_hard_row = ["Time"] + [get_value(file, 0) for file in file_names]
            best_soft_row = ["Hard"] + [get_value(file, 1) for file in file_names]
            iteration_row = ["Soft"] + [get_value(file, 2) for file in file_names]

            ws.append(best_hard_row)
            ws.append(best_soft_row)
            ws.append(iteration_row)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True)
            
            wb.save(filename)
            print(f"Results saved to '{filename}' successfully!\n")
            return
        except PermissionError:
            print(f"Permission denied: {filename} is open. Please close the file. Retrying in {WAIT_TIME} seconds...")
            time.sleep(WAIT_TIME)


def parse_last_log_line(line):
    """
    Extracts time, hard penalty, and soft penalty values from a log line.

    Parameters:
        line (str): Line from the log file.

    Returns:
        list: [time (str), hard (int), soft (int)] or None if format is invalid.
    """
    match = re.search(r"Time: ([\d:.]+), Hard: (-?\d+), Soft: (-?\d+)", line)
    if match:
        time_value = match.group(1)
        hard_value = int(match.group(2))
        soft_value = int(match.group(3))
        return [time_value, hard_value, soft_value]
    return None


def get_last_log_line(filename):
    """
    Gets the last log line from a file and parses it.

    Parameters:
        filename (str): Path to the log file.

    Returns:
        list or str: Parsed [time, hard, soft] values or error message.
    """
    try:
        with open(filename, 'rb') as f:
            f.seek(0, os.SEEK_END)
            if f.tell() == 0:
                return ""
            
            f.seek(-2, os.SEEK_END)
            while f.read(1) != b'\n':
                f.seek(-2, os.SEEK_CUR)
            return parse_last_log_line(f.readline().decode())
    except FileNotFoundError:
        return "File not found"
    except OSError:
        with open(filename, "r", encoding="utf-8") as f:
            lines = f.readlines()
            return parse_last_log_line(lines[-1]) if lines else ""
        

def profile_execution(profiler, input_file_name):
    """
    Saves the profiling statistics to a file.

    Parameters:
        profiler (cProfile.Profile): The profiler instance.
        input_file_name (str): Filename to save the profiler output.
    """
    profiler.disable()
    
    print(f"Saving profiler output...")
    
    output_dir = "profiler"
    directory_exists(output_dir)
    output_file_name = os.path.join(output_dir, input_file_name)

    s = io.StringIO()
    sortby = 'cumulative'
    ps = pstats.Stats(profiler, stream=s).sort_stats(sortby)
    ps.print_stats()

    with open(output_file_name, 'w') as f:
        f.write(s.getvalue())
    
    print(f"Profiler output saved to {output_file_name}\n")